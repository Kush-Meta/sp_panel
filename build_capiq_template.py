"""Generate a Capital IQ Pro (Excel plug-in) signal-puller workbook.

Every data cell is a live `=CIQ(...)` formula that resolves when opened in Excel
with the Capital IQ Office plug-in installed and logged in. The mnemonic and the
period for each column live in editable header cells, so if a column errors you
fix ONE cell (verify strings in the CIQ ribbon -> Formula Builder) and the whole
column updates.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

uni = json.loads(open("/tmp/uni.json").read().splitlines()[-1])

FONT = "Arial"
NAVY = "1F3864"; LBLUE = "D9E1F2"; YELL = "FFF2CC"; GREY = "F2F2F2"
hdr = Font(name=FONT, bold=True, color="FFFFFF", size=10)
lbl = Font(name=FONT, bold=True, size=9)
mn = Font(name=FONT, color="0000FF", size=9)          # blue = editable input
per = Font(name=FONT, color="0000FF", bold=True, size=9)
body = Font(name=FONT, size=9)
note = Font(name=FONT, italic=True, size=9, color="595959")
navy_fill = PatternFill("solid", fgColor=NAVY)
lblue_fill = PatternFill("solid", fgColor=LBLUE)
yell_fill = PatternFill("solid", fgColor=YELL)
grey_fill = PatternFill("solid", fgColor=GREY)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ---------------------------------------------------------------------------
# Read me
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Read me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 104
rows = [
    ("Capital IQ Pro — Revenue-forecasting signal puller", "title"),
    ("", None),
    ("What this is", "h"),
    ("A template of live =CIQ(...) formulas for the 100-company universe. It pulls the signals most likely to "
     "improve the next-quarter revenue-growth model — above all CONSENSUS ESTIMATES, which the SEC/price data "
     "cannot provide.", "p"),
    ("", None),
    ("How to use it", "h"),
    ("1.  Open in Excel with the Capital IQ Office plug-in installed and logged in (Capital IQ ribbon).", "p"),
    ("2.  Press Refresh (Capital IQ ribbon) or Ctrl+Alt+F9 to pull. First refresh can take a minute.", "p"),
    ("3.  Each data column has its MNEMONIC in row 2 and its PERIOD in row 3 (both blue = editable). "
     "If a column errors, fix that one mnemonic/period cell and the whole column updates.", "p"),
    ("4.  Add tickers on 'Universe'; the data sheets read from it, so extend formulas down to match.", "p"),
    ("", None),
    ("IMPORTANT — verify mnemonics", "hy"),
    ("Exact CIQ mnemonic spellings vary slightly by account/version (e.g. IQ_TOTAL_REV_MEDIAN_EST vs "
     "IQ_TOTAL_REV_EST_MEDIAN). Yellow header cells are best-guess strings to VERIFY once in the CIQ ribbon → "
     "Formula Builder, then they are locked in for every row. Nothing here can resolve without the plug-in.", "note"),
    ("", None),
    ("Period syntax (put in row 3 cells)", "h"),
    ("IQ_FQ = latest reported fiscal quarter · IQ_FQ-1, IQ_FQ-2 … = prior quarters · IQ_FQ+1, IQ_FQ+2 … = "
     "forward (estimate) quarters · IQ_FY / IQ_FY+1 = fiscal year / next · IQ_LTM = last twelve months · "
     "IQ_CY / IQ_CQ = calendar year / quarter.", "p"),
    ("", None),
    ("Sheets", "h"),
    ("Universe — the 100 tickers + sector, and the CIQ identifier each data sheet references.", "p"),
    ("Estimates — forward consensus revenue / EPS / EBITDA (mean, # analysts, high, low, std dev, median). "
     "THE priority block: gives a real consensus baseline and 'surprise vs consensus' target.", "p"),
    ("Actuals & Surprise — reported revenue/EPS by period + realised surprise %, and period/filing dates for the "
     "filing-date event study.", "p"),
    ("Guidance — management revenue/EPS guidance (high/low/mid).", "p"),
    ("Pricing & Valuation — market cap, TEV, close, P/E, EV/EBITDA snapshot.", "p"),
    ("", None),
    ("Identifiers", "h"),
    ("CIQ resolves plain tickers (AAPL) for US names; for ambiguity use exchange prefix (NASDAQ:AAPL, NYSE:JPM) "
     "or ticker:country (AAPL:US). BRK.B may need 'BRK.B:US'. Adjust the 'CIQ ID' column on Universe if a row fails.", "note"),
]
r = 1
for text, kind in rows:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "title":
        c.font = Font(name=FONT, bold=True, size=14, color=NAVY)
    elif kind == "h":
        c.font = Font(name=FONT, bold=True, size=11, color=NAVY)
    elif kind == "hy":
        c.font = Font(name=FONT, bold=True, size=11, color="C00000")
    elif kind == "note":
        c.font = note; c.fill = yell_fill; c.alignment = left
    else:
        c.font = body; c.alignment = left
    if kind in ("p", "note"):
        ws.row_dimensions[r].height = 30 if len(text) > 95 else 15
    r += 1

# ---------------------------------------------------------------------------
# Universe
# ---------------------------------------------------------------------------
uw = wb.create_sheet("Universe")
uw.sheet_view.showGridLines = False
for col, (w, h) in {"A": (10, "Ticker"), "B": (24, "Sector"), "C": (14, "CIQ ID")}.items():
    uw.column_dimensions[col].width = w
    cell = uw[f"{col}1"]; cell.value = h; cell.font = hdr; cell.fill = navy_fill
    cell.alignment = center; cell.border = border
uw["C1"].comment = None
for i, rec in enumerate(uni, start=2):
    a = uw.cell(row=i, column=1, value=rec["ticker"]); a.font = body; a.border = border
    b = uw.cell(row=i, column=2, value=rec["sector"]); b.font = body; b.border = border
    # CIQ identifier defaults to the ticker (editable/blue)
    c = uw.cell(row=i, column=3, value=f"=A{i}"); c.font = Font(name=FONT, color="0000FF", size=9); c.border = border
uw.freeze_panes = "A2"
N = len(uni) + 1  # last data row index

# ---------------------------------------------------------------------------
# Helper to lay out a CIQ data sheet with mnemonic/period header rows
# ---------------------------------------------------------------------------
def build_ciq_sheet(name, columns, extra_args=""):
    """columns: list of (label, mnemonic, period, verify_bool). Data = =CIQ(id, mnem, period)."""
    s = wb.create_sheet(name)
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 10
    # A: ticker header block
    s["A1"] = "Ticker"; s["A1"].font = hdr; s["A1"].fill = navy_fill; s["A1"].alignment = center; s["A1"].border = border
    s["A2"] = "mnemonic →"; s["A2"].font = lbl; s["A2"].alignment = Alignment(horizontal="right")
    s["A3"] = "period →"; s["A3"].font = lbl; s["A3"].alignment = Alignment(horizontal="right")
    for j, (label, mnem, period, verify) in enumerate(columns):
        col = j + 2
        L = get_column_letter(col)
        s.column_dimensions[L].width = 15
        hc = s.cell(row=1, column=col, value=label); hc.font = hdr; hc.fill = navy_fill
        hc.alignment = center; hc.border = border
        mc = s.cell(row=2, column=col, value=mnem); mc.font = mn; mc.alignment = center; mc.border = border
        mc.fill = yell_fill if verify else lblue_fill
        pc = s.cell(row=3, column=col, value=period); pc.font = per; pc.alignment = center; pc.border = border
        pc.fill = lblue_fill
    # data rows
    for i in range(2, N + 1):
        row = i + 2  # data starts at Excel row 4
        tc = s.cell(row=row, column=1, value=f"=Universe!C{i}")
        tc.font = body; tc.border = border
        for j in range(len(columns)):
            col = j + 2
            L = get_column_letter(col)
            args = f"$A{row},{L}$2,{L}$3" + (("," + extra_args) if extra_args else "")
            fc = s.cell(row=row, column=col, value=f"=CIQ({args})")
            fc.font = body; fc.border = border
    s.freeze_panes = "B4"
    return s

# ---------------------------------------------------------------------------
# Estimates (the priority block): forward consensus
# ---------------------------------------------------------------------------
est_cols = []
for q in ("IQ_FQ+1", "IQ_FQ+2", "IQ_FQ+3", "IQ_FQ+4"):
    est_cols += [
        (f"Rev est mean {q[-2:]}", "IQ_TOTAL_REV_EST", q, True),
        (f"Rev est # {q[-2:]}", "IQ_TOTAL_REV_NUM_EST", q, True),
        (f"Rev est high {q[-2:]}", "IQ_TOTAL_REV_HIGH_EST", q, True),
        (f"Rev est low {q[-2:]}", "IQ_TOTAL_REV_LOW_EST", q, True),
        (f"Rev est stdev {q[-2:]}", "IQ_TOTAL_REV_STDDEV_EST", q, True),
    ]
est_cols += [
    ("EPS est mean +1", "IQ_EPS_NORM_EST", "IQ_FQ+1", True),
    ("EPS est # +1", "IQ_EPS_NORM_NUM_EST", "IQ_FQ+1", True),
    ("EBITDA est +1", "IQ_EBITDA_EST", "IQ_FQ+1", True),
    ("Rev est FY+1", "IQ_TOTAL_REV_EST", "IQ_FY+1", True),
    ("Rev est FY+2", "IQ_TOTAL_REV_EST", "IQ_FY+2", True),
]
build_ciq_sheet("Estimates", est_cols)

# ---------------------------------------------------------------------------
# Actuals & Surprise (+ dates for the filing-date study)
# ---------------------------------------------------------------------------
act_cols = []
for q in ("IQ_FQ", "IQ_FQ-1", "IQ_FQ-2", "IQ_FQ-3", "IQ_FQ-4"):
    tag = q.replace("IQ_FQ", "Q").replace("+", "").replace("-0", "")
    act_cols.append((f"Revenue {tag}", "IQ_TOTAL_REV", q, False))
act_cols += [
    ("Rev surprise % Q", "IQ_REV_SURPRISE_PERCENT", "IQ_FQ", True),
    ("Rev surprise % Q-1", "IQ_REV_SURPRISE_PERCENT", "IQ_FQ-1", True),
    ("EPS actual Q", "IQ_EPS_NORM", "IQ_FQ", True),
    ("EPS surprise % Q", "IQ_EPS_SURPRISE_PERCENT", "IQ_FQ", True),
    ("Period end date Q", "IQ_PERIOD_DATE", "IQ_FQ", True),
    ("Filing date Q", "IQ_FILING_DATE", "IQ_FQ", True),
    ("Filing date Q-1", "IQ_FILING_DATE", "IQ_FQ-1", True),
]
build_ciq_sheet("Actuals & Surprise", act_cols)

# ---------------------------------------------------------------------------
# Guidance (management)
# ---------------------------------------------------------------------------
gd_cols = [
    ("Rev guid high FQ+1", "IQ_REV_GUIDANCE_HIGH", "IQ_FQ+1", True),
    ("Rev guid low FQ+1", "IQ_REV_GUIDANCE_LOW", "IQ_FQ+1", True),
    ("Rev guid high FY+1", "IQ_REV_GUIDANCE_HIGH", "IQ_FY+1", True),
    ("Rev guid low FY+1", "IQ_REV_GUIDANCE_LOW", "IQ_FY+1", True),
    ("EPS guid high FY+1", "IQ_EPS_GUIDANCE_HIGH", "IQ_FY+1", True),
    ("EPS guid low FY+1", "IQ_EPS_GUIDANCE_LOW", "IQ_FY+1", True),
]
build_ciq_sheet("Guidance", gd_cols)

# ---------------------------------------------------------------------------
# Pricing & Valuation (snapshot)
# ---------------------------------------------------------------------------
px_cols = [
    ("Close price", "IQ_CLOSEPRICE", "IQ_CY", False),
    ("Market cap", "IQ_MARKETCAP", "IQ_CY", False),
    ("TEV", "IQ_TEV", "IQ_CY", False),
    ("P/E (LTM)", "IQ_PE_EXCL", "IQ_LTM", True),
    ("EV/EBITDA (LTM)", "IQ_TEV_EBITDA", "IQ_LTM", True),
    ("EV/Revenue (LTM)", "IQ_TEV_TOTAL_REV", "IQ_LTM", True),
]
build_ciq_sheet("Pricing & Valuation", px_cols)

wb.save("capiq_signal_template.xlsx")
print("saved capiq_signal_template.xlsx  sheets:", wb.sheetnames, " tickers:", len(uni))
